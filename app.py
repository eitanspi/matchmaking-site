"""Matchmaking site — a small Flask app for managing candidates and matches.

Core features:
  * Candidates: list + filter, view, add/edit/delete, photos, VIP/"takiru" flags
  * Matches: create manually, track status, list
  * AI matches: upload a suggestions spreadsheet (name | suggestion | score ...),
    then browse the results grouped per candidate
  * Excel: export the candidate list / import (create or update) from a sheet
  * Compare: view two candidates side by side

Run locally:  python app.py   (http://localhost:5555)
"""
import os
import re
import tempfile

from flask import (Flask, flash, redirect, render_template, request,
                   send_file, url_for)
from werkzeug.utils import secure_filename

from models import (init_db, get_all_candidates, get_candidate, get_candidates_by_gender,
                    create_candidate, update_candidate, delete_candidate, candidate_exists,
                    find_candidate_by_name, set_flag, add_photo, delete_photo,
                    create_match, update_match, delete_match, get_match, get_all_matches,
                    add_ai_suggestion, get_ai_suggestions_for, get_all_ai_suggestions,
                    delete_ai_suggestion, delete_all_ai_suggestions)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
PHOTOS_DIR = os.path.join(BASE_DIR, 'static', 'photos')
os.makedirs(PHOTOS_DIR, exist_ok=True)

ALLOWED_PHOTO_EXT = {'jpg', 'jpeg', 'png', 'gif', 'webp'}

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32 MB uploads


# ---- helpers ---------------------------------------------------------------
GENDER_HE = {'male': 'זכר', 'female': 'נקבה'}


def _strip_age(name):
    """'רון כהן (28)' -> 'רון כהן' (the upload sheets append an age)."""
    return re.sub(r'\s*\(\d+\)\s*$', '', str(name).strip())


def _form_candidate(existing=None):
    def g(field, default=None):
        return request.form.get(field, default)
    return {
        'name': g('name', existing.name if existing else '') or '',
        'age': request.form.get('age', type=int) or (existing.age if existing else None),
        'gender': g('gender', existing.gender if existing else '') or None,
        'height': request.form.get('height', type=int) or (existing.height if existing else None),
        'religious_level': g('religious_level', existing.religious_level if existing else ''),
        'location': g('location', existing.location if existing else ''),
        'phone': g('phone', existing.phone if existing else ''),
        'occupation': g('occupation', existing.occupation if existing else ''),
        'description': g('description', existing.description if existing else ''),
        'looking_for': g('looking_for', existing.looking_for if existing else ''),
        'references': g('references', existing.references if existing else ''),
        'ethnicity': g('ethnicity', existing.ethnicity if existing else ''),
        'marital_status': g('marital_status', existing.marital_status if existing else ''),
    }


def _save_photos(candidate_id):
    for photo in request.files.getlist('photos'):
        if not photo or not photo.filename:
            continue
        ext = photo.filename.rsplit('.', 1)[-1].lower() if '.' in photo.filename else ''
        if ext not in ALLOWED_PHOTO_EXT:
            continue
        filename = f'{candidate_id}_{secure_filename(photo.filename)}'
        photo.save(os.path.join(PHOTOS_DIR, filename))
        add_photo(candidate_id, filename)


@app.template_filter('gender_he')
def gender_he(g):
    return GENDER_HE.get(g, '')


# ---- candidates ------------------------------------------------------------
@app.route('/')
def index():
    filters = {}
    for key in ('name', 'gender', 'religious', 'location', 'ethnicity'):
        if request.args.get(key):
            filters[key] = request.args.get(key)
    statuses = request.args.getlist('marital_status')
    if statuses:
        filters['marital_status'] = statuses if len(statuses) > 1 else statuses[0]
    for key in ('age_min', 'age_max', 'height_min', 'height_max'):
        if request.args.get(key, type=int):
            filters[key] = request.args.get(key, type=int)
    if request.args.get('vip'):
        filters['vip'] = True
    if request.args.get('takiru'):
        filters['takiru'] = True

    candidates = get_all_candidates(filters or None)
    return render_template('index.html', candidates=candidates, args=request.args)


@app.route('/candidate/<int:id>')
def candidate_detail(id):
    candidate = get_candidate(id)
    if not candidate:
        flash('מועמד/ת לא נמצא/ה', 'error')
        return redirect(url_for('index'))

    if candidate.gender in ('male', 'female'):
        opposite = 'female' if candidate.gender == 'male' else 'male'
        potentials = get_candidates_by_gender(opposite)
    else:
        potentials = [c for c in get_all_candidates() if c.id != id]

    ai_matches = []
    for s in get_ai_suggestions_for(id):
        suggested = get_candidate(s['suggested_id'])
        if suggested:
            ai_matches.append({'candidate': suggested, 'score': s.get('score'),
                               'rank': s.get('rank'), 'explanation': s.get('explanation')})

    return render_template('candidate.html', candidate=candidate, potentials=potentials,
                           partner_ids=candidate.match_partner_ids, ai_matches=ai_matches)


@app.route('/candidate/new', methods=['GET', 'POST'])
def candidate_new():
    if request.method == 'POST':
        data = _form_candidate()
        data['source'] = 'manual'
        if not data['name']:
            flash('חובה להזין שם', 'error')
            return render_template('edit_candidate.html', candidate=None)
        cid = create_candidate(data)
        _save_photos(cid)
        flash('מועמד/ת נוסף/ה בהצלחה', 'success')
        return redirect(url_for('candidate_detail', id=cid))
    return render_template('edit_candidate.html', candidate=None)


@app.route('/candidate/<int:id>/edit', methods=['GET', 'POST'])
def candidate_edit(id):
    candidate = get_candidate(id)
    if not candidate:
        flash('מועמד/ת לא נמצא/ה', 'error')
        return redirect(url_for('index'))
    if request.method == 'POST':
        del_photo = request.form.get('delete_photo', type=int)
        if del_photo:
            delete_photo(del_photo, PHOTOS_DIR)
            flash('התמונה נמחקה', 'info')
            return redirect(url_for('candidate_edit', id=id))
        update_candidate(id, _form_candidate(candidate))
        _save_photos(id)
        flash('הפרופיל עודכן', 'success')
        return redirect(url_for('candidate_detail', id=id))
    return render_template('edit_candidate.html', candidate=candidate)


@app.route('/candidate/<int:id>/delete', methods=['POST'])
def candidate_delete(id):
    delete_candidate(id)
    flash('המועמד/ת נמחק/ה', 'info')
    return redirect(url_for('index'))


@app.route('/candidate/<int:id>/toggle/<flag>', methods=['POST'])
def candidate_toggle(id, flag):
    if flag not in ('vip', 'takiru'):
        flash('דגל לא תקין', 'error')
        return redirect(request.referrer or url_for('index'))
    candidate = get_candidate(id)
    if candidate:
        set_flag(id, flag, not bool(getattr(candidate, flag)))
    return redirect(request.referrer or url_for('candidate_detail', id=id))


# ---- matches ---------------------------------------------------------------
@app.route('/match/create', methods=['POST'])
def match_create():
    a_id = request.form.get('candidate_a_id', type=int)
    b_id = request.form.get('candidate_b_id', type=int)
    notes = request.form.get('notes', '')
    if not a_id or not b_id or a_id == b_id:
        flash('בחירה לא תקינה', 'error')
        return redirect(request.referrer or url_for('index'))
    if create_match(a_id, b_id, notes) is None:
        flash('הצעת שידוך כבר קיימת', 'warning')
    else:
        flash('הצעת שידוך נוצרה', 'success')
    return redirect(url_for('candidate_detail', id=a_id))


@app.route('/match/<int:id>/update', methods=['POST'])
def match_update(id):
    update_match(id, status=request.form.get('status'), notes=request.form.get('notes'))
    flash('הסטטוס עודכן', 'success')
    return redirect(request.referrer or url_for('matches'))


@app.route('/match/<int:id>/delete', methods=['POST'])
def match_delete_route(id):
    delete_match(id)
    flash('ההצעה נמחקה', 'info')
    return redirect(request.referrer or url_for('matches'))


@app.route('/matches')
def matches():
    status = request.args.get('status')
    return render_template('matches.html', matches=get_all_matches(status), status=status)


# ---- compare ---------------------------------------------------------------
@app.route('/compare')
def compare():
    a = get_candidate(request.args.get('a', type=int))
    b = get_candidate(request.args.get('b', type=int))
    return render_template('compare.html', candidate_a=a, candidate_b=b)


# ---- AI matches ------------------------------------------------------------
@app.route('/ai-matches')
def ai_matches():
    sort = request.args.get('sort', 'score')
    grouped = {}
    for s in get_all_ai_suggestions():
        cid = s['candidate_id']
        grouped.setdefault(cid, {
            'candidate_id': cid, 'candidate_name': s['candidate_name'],
            'candidate_age': s['candidate_age'], 'candidate_gender': s['candidate_gender'],
            'suggestions': [],
        })['suggestions'].append(s)
    groups = list(grouped.values())
    if sort == 'name':
        groups.sort(key=lambda g: g['candidate_name'] or '')
    elif sort == 'age':
        groups.sort(key=lambda g: g['candidate_age'] or 99)
    else:
        groups.sort(key=lambda g: max((s['score'] or 0) for s in g['suggestions']), reverse=True)
    return render_template('ai_matches.html', groups=groups, sort=sort)


@app.route('/ai-matches/upload', methods=['GET', 'POST'])
def ai_matches_upload():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('יש להעלות קובץ Excel (.xlsx)', 'error')
            return redirect(url_for('ai_matches_upload'))
        import openpyxl
        ws = openpyxl.load_workbook(f).active
        added, not_found = 0, []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            main = find_candidate_by_name(_strip_age(row[0]))
            if not main:
                not_found.append(_strip_age(row[0]))
                continue
            # suggestions are (name, score) pairs across the remaining columns
            i, rank = 1, 1
            while i < len(row) and row[i]:
                sug = find_candidate_by_name(_strip_age(row[i]))
                score = float(row[i + 1]) if i + 1 < len(row) and row[i + 1] else 0.0
                if sug:
                    add_ai_suggestion(main.id, sug.id, rank, score)
                    added += 1
                else:
                    not_found.append(_strip_age(row[i]))
                rank += 1
                i += 2
        msg = f'נוספו {added} הצעות AI.'
        if not_found:
            msg += ' לא נמצאו: ' + ', '.join(sorted(set(not_found))[:10])
        flash(msg, 'success' if added else 'warning')
        return redirect(url_for('ai_matches'))
    return render_template('ai_matches_upload.html')


@app.route('/ai-suggestion/<int:id>/delete', methods=['POST'])
def ai_suggestion_delete(id):
    delete_ai_suggestion(id)
    return redirect(request.referrer or url_for('ai_matches'))


@app.route('/ai-matches/delete-all', methods=['POST'])
def ai_matches_delete_all():
    count = delete_all_ai_suggestions()
    flash(f'נמחקו {count} הצעות', 'info')
    return redirect(url_for('ai_matches'))


# ---- Excel import / export -------------------------------------------------
EXCEL_COLUMNS = [
    ('ID', 'id', 8), ('שם', 'name', 25), ('גיל', 'age', 8), ('מין', 'gender', 10),
    ('גובה', 'height', 8), ('רמה דתית', 'religious_level', 18), ('מיקום', 'location', 18),
    ('טלפון', 'phone', 15), ('עיסוק', 'occupation', 25), ('סטטוס', 'marital_status', 12),
    ('עדה', 'ethnicity', 15), ('VIP', 'vip', 6), ('תכירו', 'takiru', 6),
    ('תיאור', 'description', 50), ('מחפש/ת', 'looking_for', 40), ('ממליצים', 'references', 30),
]


@app.route('/excel')
def excel_page():
    return render_template('excel.html')


@app.route('/excel/download')
def excel_download():
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'מועמדים'
    ws.sheet_view.rightToLeft = True
    hf = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='4472C4')
    border = Border(*[Side(style='thin')] * 4)
    for ci, (header, _, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font, cell.fill = hf, fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width
    for ri, c in enumerate(get_all_candidates(), 2):
        for ci, (_, field, _) in enumerate(EXCEL_COLUMNS, 1):
            if field == 'gender':
                val = GENDER_HE.get(c.gender, '')
            elif field in ('vip', 'takiru'):
                val = 'כן' if getattr(c, field) else ''
            else:
                val = getattr(c, field) or ''
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'
    path = os.path.join(tempfile.gettempdir(), 'candidates.xlsx')
    wb.save(path)
    return send_file(path, as_attachment=True, download_name='candidates.xlsx')


@app.route('/excel/upload', methods=['POST'])
def excel_upload():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('יש להעלות קובץ Excel (.xlsx)', 'error')
        return redirect(url_for('excel_page'))
    import openpyxl
    ws = openpyxl.load_workbook(file).active
    headers = {c.value.strip(): i for i, c in enumerate(ws[1], 1) if c.value}
    field_map = {h: f for h, f, _ in EXCEL_COLUMNS if f not in ('id', 'vip', 'takiru')}
    created = updated = 0
    for row in ws.iter_rows(min_row=2):
        def cell(header):
            col = headers.get(header)
            return row[col - 1].value if col else None
        data = {}
        for heb, field in field_map.items():
            val = cell(heb)
            if field == 'gender':
                val = {'זכר': 'male', 'נקבה': 'female'}.get(val)
            elif field in ('age', 'height'):
                val = int(val) if val else None
            data[field] = val if val not in ('', None) else None
        if not data.get('name'):
            continue
        is_vip = str(cell('VIP') or '').strip() in ('כן', '1', 'True', 'true', 'yes')
        is_takiru = str(cell('תכירו') or '').strip() in ('כן', '1', 'True', 'true', 'yes')
        row_id = cell('ID')
        existing = get_candidate(int(row_id)) if row_id else None
        if existing:
            update_candidate(existing.id, data)
            set_flag(existing.id, 'vip', is_vip)
            set_flag(existing.id, 'takiru', is_takiru)
            updated += 1
        else:
            data['source'] = 'excel'
            cid = create_candidate(data)
            set_flag(cid, 'vip', is_vip)
            set_flag(cid, 'takiru', is_takiru)
            created += 1
    flash(f'עודכנו {updated} מועמדים, נוצרו {created} חדשים', 'success')
    return redirect(url_for('excel_page'))


init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5555)
